import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

# Define a function to store data in an Excel file
def save_to_excel(data, filename='form_data.xlsx'):
    df = pd.DataFrame([data])

    if not os.path.exists(filename):
        # Create a new Excel file if it doesn't exist
        df.to_excel(filename, index=False)
    else:
        try:
            # Load the existing workbook
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book

            # Determine the last row in the existing sheet
            startrow = book['Sheet1'].max_row

            # Append the new data
            df.to_excel(writer, index=False, header=False, startrow=startrow)
            writer.save()
        except Exception as e:
            st.error(f"An error occurred: {e}")

# Function to load and display Excel data
def load_and_display_excel(filename='form_data.xlsx'):
    if os.path.exists(filename):
        df = pd.read_excel(filename)
        st.dataframe(df)  # Display the DataFrame in Streamlit
    else:
        st.warning("No data found. Please submit the form first.")

# Streamlit App UI
st.title("User Input Form")

# Create a form
with st.form(key="user_form"):
    # 1st Input: Text box
    name = st.text_input("Enter your name")

    # 2nd Input: Radio button
    gender = st.radio("Select your gender", ("Male", "Female", "Other"))

    # Submit button
    submit_button = st.form_submit_button(label="Submit")
    show=st.form_submit_button(label="Show data")

# When form is submitted
if submit_button:
    if name and gender:
        # Prepare data for Excel
        data = {"Name": name, "Gender": gender}
        save_to_excel(data)
        

        # Confirmation message
        st.success(f"Data saved! Name: {name}, Gender: {gender}")

        # Display the updated Excel data
        load_and_display_excel()
    else:
        st.error("Please provide both inputs.")
if show:
    load_and_display_excel()
